from flask import Flask, render_template, request, send_file
import pandas as pd
import time
import io
import concurrent.futures
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)

# Define color fills
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')


def fetch_delivery_info(asin, pincode):
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--remote-debugging-port=9222")
    options.add_argument("--disable-extensions")
    options.add_argument("--incognito")  # 🔥 This enables Incognito Mode

    driver = webdriver.Chrome(service=Service("/usr/local/bin/chromedriver"), options=options)

    result = {
        "ASIN": asin,
        "PIN Code": pincode,
        "In Stock": "Error",
        "Delivery Info": "Could not fetch"
    }

    try:
        print(f"[INFO] Opening Amazon Homepage for PIN: {pincode}")
        driver.get("https://www.amazon.in/")
        time.sleep(2)

        # Set PIN Code
        try:
            driver.find_element(By.ID, "nav-global-location-popover-link").click()
            time.sleep(2)

            pin_input = driver.find_element(By.ID, "GLUXZipUpdateInput")
            pin_input.clear()
            pin_input.send_keys(pincode)
            time.sleep(1)

            driver.find_element(By.XPATH, "//span[@id='GLUXZipUpdate']//input").click()
            time.sleep(3)

            try:
                driver.find_element(By.NAME, "glowDoneButton").click()
                time.sleep(2)
            except Exception as e:
                pass
        except Exception as e:
            print(f"[ERROR] Failed to set PIN for {asin} - {pincode}: {e}")

        # Now visit the product page
        print(f"[INFO] Fetching Product Page for ASIN: {asin}")
        url = f"https://www.amazon.in/dp/{asin}"
        driver.get(url)
        time.sleep(3)

        # Try to get delivery message
        try:
            delivery_element = driver.find_element(By.ID, "delivery-message")
            delivery_text = delivery_element.text.strip()
        except:
            try:
                delivery_element = driver.find_element(By.XPATH, "//div[@id='mir-layout-DELIVERY_BLOCK-slot-PRIMARY_DELIVERY_MESSAGE_LARGE']")
                delivery_text = delivery_element.text.strip()
            except:
                delivery_text = "Unavailable"

        # 🛡️ **Improved Logic to Check Delivery Status**
        if any(keyword in delivery_text.lower() for keyword in [
            "not deliverable",
            "cannot be delivered",
            "unavailable",
            "currently out of stock",
            "we don't deliver to this address"
        ]):
            result["In Stock"] = "No"
        else:
            result["In Stock"] = "Yes"

        result["Delivery Info"] = delivery_text

    except Exception as e:
        result["Delivery Info"] = f"Error: {str(e)}"

    driver.quit()
    return result


def check_availability(asins, pincodes):
    combos = [(asin, pin) for asin in asins for pin in pincodes]
    results = []

    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
        futures = [executor.submit(fetch_delivery_info, asin, pin) for asin, pin in combos]
        for future in concurrent.futures.as_completed(futures):
            results.append(future.result())

    return pd.DataFrame(results)


def colorize_excel(df, output):
    """
    Adds color coding to the Excel sheet:
    - Green for In Stock
    - Red for Not In Stock
    - Yellow for Error
    """
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Results")
        workbook = writer.book
        sheet = workbook.active

        for row in range(2, sheet.max_row + 1):
            status = sheet[f'C{row}'].value
            if status == "Yes":
                for col in 'ABC':
                    sheet[f'{col}{row}'].fill = GREEN_FILL
            elif status == "No":
                for col in 'ABC':
                    sheet[f'{col}{row}'].fill = RED_FILL
            else:
                for col in 'ABC':
                    sheet[f'{col}{row}'].fill = YELLOW_FILL


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        asin_input = request.form["asins"]
        pincode_input = request.form["pincodes"]

        asin_list = [x.strip() for x in asin_input.splitlines() if x.strip()]
        pincode_list = [x.strip() for x in pincode_input.splitlines() if x.strip()]

        df = check_availability(asin_list, pincode_list)

        # Save to Excel in memory with color coding
        output = io.BytesIO()
        colorize_excel(df, output)
        output.seek(0)

        return send_file(output, as_attachment=True, download_name="availability_report.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    return render_template("index.html")


if __name__ == "__main__":
    app.run(debug=True)
